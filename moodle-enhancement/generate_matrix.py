"""
Generate Airpay Academy Access Matrix as a proper Excel file.
Includes ALL Moodle + BizLMS functionality + planned features.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Access Matrix"

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='0066A7', end_color='0066A7', fill_type='solid')
cat_font = Font(name='Calibri', bold=True, size=10, color='0066A7')
cat_fill = PatternFill(start_color='E8F2F9', end_color='E8F2F9', fill_type='solid')
yes_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
no_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
plan_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
fix_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 48
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 16
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 18

roles = ['Super Admin', 'L&D Admin', 'Manager/HRBP', 'Employee', 'External Learner', 'Trainer', 'Guest', 'Status']

# Helper
Y = 'YES'
N = 'NO'
P = 'PLANNED'
F = 'NEEDS FIX'
W = 'WORKS'

data = [
    # Title row
    ['AIRPAY ACADEMY — COMPLETE ACCESS MATRIX', '', '', '', '', '', '', '', '', ''],
    ['Generated: 2026-04-04 | Owner: Nitin Rajput | 7 User Types × 100+ Features', '', '', '', '', '', '', '', '', ''],
    ['', '', '', '', '', '', '', '', '', ''],
    # Header
    ['Category', 'Feature / Functionality', *roles],

    # ═══ DASHBOARD ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['DASHBOARD', '── Admin Dashboard Sections ──', '', '', '', '', '', '', '', ''],
    ['Dashboard', 'KPI Tiles (total users, courses, enrolments, completion rate)', Y, Y, N, N, N, N, N, W],
    ['Dashboard', 'Quick Navigation cards (with live stats)', Y, Y, N, N, N, N, N, W],
    ['Dashboard', 'Enrolment Trend bar chart (Chart.js, 6 months)', Y, Y, N, N, N, N, N, W],
    ['Dashboard', 'Course Distribution doughnut chart', Y, Y, N, N, N, N, N, W],
    ['Dashboard', 'System Health (cron, disk, PHP, Moodle version)', Y, N, N, N, N, N, N, F],
    ['Dashboard', 'User Analytics (logins today/week, new users, inactive 30d+)', Y, Y, N, N, N, N, N, W],
    ['Dashboard', '', '', '', '', '', '', '', '', ''],
    ['Dashboard', '── Manager Dashboard Sections ──', '', '', '', '', '', '', '', ''],
    ['Dashboard', 'Team KPIs (members, enrolments, completions, rate)', N, N, Y, N, N, N, N, W],
    ['Dashboard', 'Team Compliance Table (name, enrolled, completed, pending, last active)', N, N, Y, N, N, N, N, W],
    ['Dashboard', '', '', '', '', '', '', '', '', ''],
    ['Dashboard', '── Learner Dashboard Sections ──', '', '', '', '', '', '', '', ''],
    ['Dashboard', 'Welcome Banner (greeting + Browse Courses / My Profile)', N, N, Y, Y, Y, Y, N, W],
    ['Dashboard', 'Stat Cards (enrolled, in-progress, completed, certificates)', N, N, Y, Y, Y, Y, N, W],
    ['Dashboard', 'Continue Learning (course cards with progress bars)', N, N, Y, Y, Y, Y, N, W],
    ['Dashboard', 'Upcoming Deadlines (courses with due dates, urgent flags)', N, N, Y, Y, Y, N, N, W],
    ['Dashboard', 'Recent Achievements (certificates with codes + dates)', N, N, Y, Y, Y, N, N, W],
    ['Dashboard', 'Activity Timeline (enrolments, completions, quiz submits)', N, N, Y, Y, Y, N, N, W],
    ['Dashboard', 'Recommended Courses (category-match unenrolled courses)', N, N, Y, Y, Y, N, N, W],

    # ═══ NAVBAR ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['NAVBAR', '── Navigation Bar ──', '', '', '', '', '', '', '', ''],
    ['Navbar', 'Dashboard pill (always visible)', Y, Y, Y, Y, Y, Y, Y, W],
    ['Navbar', 'My Courses pill', N, N, Y, Y, Y, Y, N, F],
    ['Navbar', 'Catalog pill', N, N, Y, Y, Y, N, Y, F],
    ['Navbar', 'Profile pill', N, N, Y, Y, Y, Y, N, F],
    ['Navbar', 'Search bar', Y, Y, Y, Y, Y, Y, Y, W],
    ['Navbar', 'Quick Access popover (BizLMS admin menu)', Y, Y, Y, N, N, Y, N, W],
    ['Navbar', 'Notifications bell', Y, Y, Y, Y, Y, Y, N, W],
    ['Navbar', 'Messaging', Y, Y, Y, Y, Y, Y, N, W],
    ['Navbar', 'Shopping Cart', N, N, N, Y, Y, N, N, W],
    ['Navbar', 'Dark mode toggle', Y, Y, Y, Y, Y, Y, Y, W],
    ['Navbar', 'Edit mode toggle', Y, Y, N, N, N, N, N, W],
    ['Navbar', 'User menu dropdown (logout, profile, grades)', Y, Y, Y, Y, Y, Y, N, F],

    # ═══ QUICK ACCESS MENU ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['QUICK ACCESS', '── Quick Access Menu Items ──', '', '', '', '', '', '', '', ''],
    ['Quick Access', 'Dashboard', Y, Y, Y, N, N, N, N, W],
    ['Quick Access', 'Company Structure (costcenter org tree)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Users (create, edit, bulk upload, sync)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Groups', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Search / Catalog', Y, Y, Y, N, N, N, N, W],
    ['Quick Access', 'View Transactions (payments)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Categories (custom course categories)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Courses (create, edit, enrol, bulk upload)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Forum (create, moderate)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Online Exams (create, edit, results)', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Classrooms (ILT sessions, attendance)', Y, Y, Y, N, N, Y, N, W],
    ['Quick Access', 'Manage Learning Paths', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Programs', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Feedbacks / Evaluations', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Manage Skills Repository', Y, Y, N, N, N, N, N, W],
    ['Quick Access', 'Notifications (create, manage)', Y, Y, Y, N, N, N, N, W],
    ['Quick Access', 'Manage Requests (approve/deny)', Y, Y, Y, N, N, N, N, W],
    ['Quick Access', 'Analytics / LearnerScript Reports', Y, Y, Y, N, N, N, N, W],
    ['Quick Access', 'Trainer Dashboard', N, N, N, N, N, Y, N, W],
    ['Quick Access', 'Administration (Moodle Site Admin)', Y, N, N, N, N, N, N, W],

    # ═══ BIZLMS CORE MODULES ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['BIZLMS MODULES', '── BizLMS Core Functionality ──', '', '', '', '', '', '', '', ''],
    ['Courses', 'Create / edit e-learning courses', Y, Y, N, N, N, N, N, W],
    ['Courses', 'Enrol / unenrol users in courses', Y, Y, N, N, N, N, N, W],
    ['Courses', 'Bulk upload courses (CSV)', Y, Y, N, N, N, N, N, W],
    ['Courses', 'View course as learner', Y, Y, Y, Y, Y, Y, N, W],
    ['Courses', 'View course completion reports', Y, Y, Y, N, N, N, N, W],
    ['Courses', 'Export course data', Y, Y, N, N, N, N, N, W],
    ['Classrooms', 'Create / edit classroom sessions (ILT)', Y, Y, N, N, N, N, N, W],
    ['Classrooms', 'Assign trainers to sessions', Y, Y, N, N, N, N, N, W],
    ['Classrooms', 'Take session attendance', Y, Y, N, N, N, Y, N, W],
    ['Classrooms', 'View classroom as trainer', N, N, N, N, N, Y, N, W],
    ['Classrooms', 'Enrol in classroom session', N, N, N, Y, Y, N, N, W],
    ['Online Exams', 'Create / edit online exams', Y, Y, N, N, N, N, N, W],
    ['Online Exams', 'View exam results / grades', Y, Y, N, N, N, N, N, W],
    ['Online Exams', 'Take an online exam', N, N, N, Y, Y, N, N, W],
    ['Programs', 'Create / edit programs (multi-course bundles)', Y, Y, N, N, N, N, N, W],
    ['Programs', 'Enrol in a program', N, N, N, Y, Y, N, N, W],
    ['Learning Plans', 'Create / edit learning paths', Y, Y, N, N, N, N, N, W],
    ['Learning Plans', 'Assign learning paths to users', Y, Y, N, N, N, N, N, W],
    ['Users', 'Create / edit users', Y, Y, N, N, N, N, N, W],
    ['Users', 'Bulk upload users (HRMS sync)', Y, Y, N, N, N, N, N, W],
    ['Users', 'Bulk status change (activate/suspend)', Y, Y, N, N, N, N, N, W],
    ['Users', 'View user profiles', Y, Y, Y, Y, Y, Y, N, W],
    ['Costcenter', 'Manage organisations / departments', Y, Y, N, N, N, N, N, W],
    ['Costcenter', 'Assign managers to departments', Y, Y, N, N, N, N, N, W],
    ['Costcenter', 'Multi-tenant branding (costcenter schemes)', Y, Y, N, N, N, N, N, W],
    ['Ratings', 'Rate courses', N, N, Y, Y, Y, N, N, W],
    ['Requests', 'Submit learning request', N, N, Y, Y, N, N, N, W],
    ['Requests', 'Approve / deny requests', Y, Y, Y, N, N, N, N, W],
    ['Forum', 'Create / moderate forums', Y, Y, N, N, N, N, N, W],
    ['Forum', 'Participate in forums', N, N, Y, Y, Y, N, N, W],
    ['Notifications', 'Create / send notifications', Y, Y, N, N, N, N, N, W],
    ['Notifications', 'Receive notifications', Y, Y, Y, Y, Y, Y, N, W],
    ['MyTeam', 'Approve team requests', N, N, Y, N, N, N, N, W],
    ['MyTeam', 'View team learning progress', N, N, Y, N, N, N, N, W],
    ['Cart', 'Purchase courses', N, N, N, Y, Y, N, N, W],
    ['Cart', 'View purchase history', N, N, N, Y, Y, N, N, W],
    ['Cart', 'Cashier / manage payments', Y, Y, N, N, N, N, N, W],

    # ═══ MOODLE CORE ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['MOODLE CORE', '── Moodle Native Features ──', '', '', '', '', '', '', '', ''],
    ['Moodle', 'Site Administration (/admin/)', Y, N, N, N, N, N, N, W],
    ['Moodle', 'Browse users (/admin/user.php)', Y, Y, N, N, N, N, N, W],
    ['Moodle', 'Manage plugins', Y, N, N, N, N, N, N, W],
    ['Moodle', 'Theme settings', Y, N, N, N, N, N, N, W],
    ['Moodle', 'SCORM player (course content)', Y, Y, Y, Y, Y, Y, N, W],
    ['Moodle', 'Quiz / assessment engine', Y, Y, Y, Y, Y, Y, N, W],
    ['Moodle', 'Gradebook', Y, Y, Y, Y, Y, Y, N, W],
    ['Moodle', 'Calendar', Y, Y, Y, Y, Y, Y, N, W],
    ['Moodle', 'Messaging', Y, Y, Y, Y, Y, Y, N, W],
    ['Moodle', 'Badges (issue/view)', Y, Y, N, Y, Y, N, N, W],
    ['Moodle', 'Certificates (tool_certificate)', Y, Y, N, Y, Y, N, N, W],
    ['Moodle', 'File management', Y, Y, Y, Y, Y, Y, N, W],
    ['Moodle', 'H5P interactive content', Y, Y, Y, Y, Y, Y, N, 'NOT ENABLED'],
    ['Moodle', 'AI subsystem', Y, N, N, N, N, N, N, 'NOT ENABLED'],

    # ═══ FOOTER ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['FOOTER', '── Footer Sections ──', '', '', '', '', '', '', '', ''],
    ['Footer', 'Full 4-column footer (Brand + Learn + Support)', N, N, N, Y, Y, Y, Y, W],
    ['Footer', 'Minimal footer (copyright only)', Y, Y, Y, N, N, N, N, W],
    ['Footer', 'Made in India / Digital India badge', Y, Y, Y, Y, Y, Y, Y, W],

    # ═══ PAGE ACCESS ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['PAGE ACCESS', '── Page-Level Access ──', '', '', '', '', '', '', '', ''],
    ['Pages', 'Dashboard (/my/)', Y, Y, Y, Y, Y, Y, N, W],
    ['Pages', 'Manage Users (/local/users/)', Y, Y, N, N, N, N, N, W],
    ['Pages', 'Manage Courses (/local/courses/)', Y, Y, N, N, N, N, N, W],
    ['Pages', 'Course Catalog (/local/search/allcourses.php)', Y, Y, Y, Y, Y, Y, N, W],
    ['Pages', 'Course Detail (/local/search/coursedetails.php)', Y, Y, Y, Y, Y, Y, N, W],
    ['Pages', 'My Profile (/local/users/profile.php)', Y, Y, Y, Y, Y, Y, N, W],
    ['Pages', 'Classrooms (/local/classroom/)', Y, Y, Y, N, N, Y, N, W],
    ['Pages', 'Online Exams (/local/onlineexams/)', Y, Y, N, N, N, N, N, W],
    ['Pages', 'Programs (/local/program/)', Y, Y, N, N, N, N, N, W],
    ['Pages', 'Learning Paths (/local/learningplan/)', Y, Y, N, N, N, N, N, W],
    ['Pages', 'Reports - LearnerScript (/blocks/learnerscript/)', Y, Y, Y, N, N, N, N, W],
    ['Pages', 'Site Admin (/admin/)', Y, N, N, N, N, N, N, W],
    ['Pages', 'Privacy Policy (/local/airpay_pages/?page=privacy)', Y, Y, Y, Y, Y, Y, Y, W],
    ['Pages', 'Terms of Use (/local/airpay_pages/?page=terms)', Y, Y, Y, Y, Y, Y, Y, W],
    ['Pages', 'Help Center (/local/airpay_pages/?page=help)', Y, Y, Y, Y, Y, Y, Y, W],
    ['Pages', 'Contact Us (/local/airpay_pages/?page=contact)', Y, Y, Y, Y, Y, Y, Y, W],

    # ═══ PLANNED FEATURES (Phase 6C/6D) ═══
    ['', '', '', '', '', '', '', '', '', ''],
    ['PLANNED', '── Phase 6C: Enhanced Capabilities ──', '', '', '', '', '', '', '', ''],
    ['Phase 6C', 'Compliance Dashboard (mandatory course matrix, RAG status)', Y, Y, Y, N, N, N, N, P],
    ['Phase 6C', 'Social Learning Feed (peer completions, activity feed)', N, N, N, Y, Y, N, N, P],
    ['Phase 6C', 'Gamification — XP / Leaderboards (block_xp)', N, N, N, Y, Y, N, N, P],
    ['Phase 6C', 'AI Quiz Generation (qbank_genai)', Y, Y, N, N, N, N, N, P],
    ['Phase 6C', 'Proctored Exams (webcam monitoring)', Y, Y, N, Y, Y, N, N, P],
    ['Phase 6C', 'Exam Portal Layout (distraction-free)', N, N, N, Y, Y, N, N, P],
    ['Phase 6C', 'HRMS Sync API (Keka real-time sync)', Y, Y, N, N, N, N, N, P],
    ['Phase 6C', 'Manager Approval Inbox (one-click approve/reject)', N, N, Y, N, N, N, N, P],
    ['Phase 6C', 'PWA Manifest + Install Prompt', N, N, N, Y, Y, N, N, P],
    ['Phase 6C', 'Push Notifications (Firebase web push)', N, N, N, Y, Y, N, N, P],
    ['Phase 6C', 'QR Attendance (classroom session check-in)', N, N, N, Y, N, Y, N, P],
    ['Phase 6C', 'Certificate Gallery (download cards, LinkedIn share)', N, N, N, Y, Y, N, N, P],
    ['', '', '', '', '', '', '', '', '', ''],
    ['PLANNED', '── Phase 6D: Analytics & Intelligence ──', '', '', '', '', '', '', '', ''],
    ['Phase 6D', 'L&D Command Center (executive KPIs, trend charts)', Y, Y, N, N, N, N, N, P],
    ['Phase 6D', 'Learner Insights ("Top 20% in your dept", skill strengths)', N, N, N, Y, Y, N, N, P],
    ['Phase 6D', 'AI Course Recommendations (ML-based, skill-gap driven)', N, N, N, Y, Y, N, N, P],
    ['Phase 6D', 'Predictive Compliance Alerts ("15 will miss AML deadline")', Y, Y, Y, N, N, N, N, P],
    ['Phase 6D', 'Training Effectiveness Reports (Kirkpatrick L1/L2)', Y, Y, Y, N, N, N, N, P],
    ['Phase 6D', 'Manager Team Analytics (skill gap heatmap, velocity)', N, N, Y, N, N, N, N, P],
    ['Phase 6D', 'Content Effectiveness Scoring (completion rate, drop-off)', Y, Y, N, N, N, N, N, P],
    ['Phase 6D', 'Microsoft Teams Integration (notifications)', Y, Y, Y, Y, Y, Y, N, P],
    ['', '', '', '', '', '', '', '', '', ''],
    ['PLANNED', '── Workstream B: SENTIENTIA ──', '', '', '', '', '', '', '', ''],
    ['SENTIENTIA', 'SOP Parser (PDF to structured JSON)', Y, Y, N, N, N, N, N, P],
    ['SENTIENTIA', 'Narration Generator (summary to script)', Y, Y, N, N, N, N, N, P],
    ['SENTIENTIA', 'Slides Generator (narration to slides)', Y, Y, N, N, N, N, N, P],
    ['SENTIENTIA', 'Voice Generator (ElevenLabs TTS)', Y, Y, N, N, N, N, N, P],
    ['SENTIENTIA', 'SCORM Packager (slides + audio to SCORM 1.2)', Y, Y, N, N, N, N, N, P],
    ['SENTIENTIA', 'Moodle Upload (SCORM to course)', Y, Y, N, N, N, N, N, P],
    ['', '', '', '', '', '', '', '', '', ''],
    ['PLANNED', '── Workstream C: Knowledge Automation ──', '', '', '', '', '', '', '', ''],
    ['Workstream C', 'Microsoft 365 Integration (Azure AD SSO)', Y, Y, Y, Y, Y, N, N, P],
    ['Workstream C', 'SharePoint Content Sync', Y, Y, N, N, N, N, N, P],
    ['Workstream C', 'Teams Channel Notifications', Y, Y, Y, Y, Y, Y, N, P],
]

# Write data
for row_idx, row_data in enumerate(data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center if col_idx > 2 else left_wrap
        cell.font = Font(name='Calibri', size=10)

        # Apply styles
        if row_idx == 4:  # Header row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        elif row_idx <= 3:  # Title
            cell.font = Font(name='Calibri', bold=True, size=12 if row_idx == 1 else 10, color='0066A7')
            cell.border = Border()
        elif value == Y:
            cell.fill = yes_fill
            cell.font = Font(name='Calibri', size=10, bold=True, color='065F46')
        elif value == N:
            cell.fill = no_fill
            cell.font = Font(name='Calibri', size=10, color='991B1B')
        elif value == P:
            cell.fill = plan_fill
            cell.font = Font(name='Calibri', size=10, bold=True, color='92400E')
        elif value == F:
            cell.fill = fix_fill
            cell.font = Font(name='Calibri', size=10, bold=True, color='1E40AF')
        elif value == W:
            cell.fill = yes_fill
            cell.font = Font(name='Calibri', size=10, color='065F46')

        # Category section headers
        if col_idx == 1 and value and value.isupper() and row_idx > 4:
            cell.font = Font(name='Calibri', bold=True, size=10, color='0066A7')
            cell.fill = cat_fill
            # Apply cat_fill to entire row
            for c in range(1, 11):
                ws.cell(row=row_idx, column=c).fill = cat_fill

# Freeze panes
ws.freeze_panes = 'C5'

# Auto-filter
ws.auto_filter.ref = f'A4:J{ws.max_row}'

# Save
filepath = r'D:\Claude Local\airpay-ld-os\moodle-enhancement\Airpay-Academy-Access-Matrix.xlsx'
wb.save(filepath)
print(f'Saved to: {filepath}')
print(f'Rows: {ws.max_row}, Columns: {ws.max_column}')
