"""
Create Excel sheet documenting all Moodle plugin upgrades during production data import.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

OUTPUT = r"D:\Claude Local\airpay-ld-os\moodle-enhancement\Production-Import-Upgrade-Log.xlsx"

# Header styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="0066A7", end_color="0066A7", fill_type="solid")
install_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
upgrade_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
delete_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
missing_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
custom_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# All plugins from the upgrade page
plugins = [
    # Activity modules
    ("Activity modules", "Assignment", "/mod/assign", "2022112801", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "BigBlueButton", "/mod/bigbluebuttonbn", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Book", "/mod/book", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Chat", "/mod/chat", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Choice", "/mod/choice", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Database", "/mod/data", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Feedback", "/mod/feedback", "2022112800", "2024100701", "Standard", "To be upgraded"),
    ("Activity modules", "Folder", "/mod/folder", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Forum", "/mod/forum", "2022112801", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Glossary", "/mod/glossary", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "H5P", "/mod/h5pactivity", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "IMS content package", "/mod/imscp", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Text and media area", "/mod/label", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Lesson", "/mod/lesson", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "External tool", "/mod/lti", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Page", "/mod/page", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Quiz", "/mod/quiz", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "File", "/mod/resource", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "SCORM package", "/mod/scorm", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Subsection", "/mod/subsection", "", "2024100700", "Standard", "To be installed"),
    ("Activity modules", "Survey", "/mod/survey", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "URL", "/mod/url", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Wiki", "/mod/wiki", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "Workshop", "/mod/workshop", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Activity modules", "assignment (legacy)", "", "2022112800", "", "Standard", "To be deleted"),
    # Blocks
    ("Blocks", "Accessibility review", "/blocks/accessreview", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Blocks", "Activities", "/blocks/activity_modules", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Blocks", "Airpay Compliance Dashboard", "/blocks/airpay_compliance", "", "2026040500", "Additional (Custom)", "To be installed"),
    ("Blocks", "Calendar", "/blocks/calendar_month", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Blocks", "Course completion status", "/blocks/completionstatus", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Blocks", "Course overview", "/blocks/myoverview", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Blocks", "Timeline", "/blocks/timeline", "2022112800", "2024100700", "Standard", "To be upgraded"),
    # Question types
    ("Question types", "Ordering", "/question/type/ordering", "", "2024100700", "Standard", "To be installed"),
    ("Question types", "Multiple choice", "/question/type/multichoice", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Question types", "Short answer", "/question/type/shortanswer", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Question types", "True/False", "/question/type/truefalse", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Question types", "Essay", "/question/type/essay", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Question types", "Calculated", "/question/type/calculated", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Question types", "Matching", "/question/type/match", "2022112800", "2024100700", "Standard", "To be upgraded"),
    # Filters
    ("Text filters", "Code highlighter", "/filter/codehighlighter", "", "2024100700", "Standard", "To be installed"),
    ("Text filters", "MathJax", "/filter/mathjaxloader", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Text filters", "tidy (legacy)", "", "2022112800", "", "Standard", "To be deleted"),
    # Editors
    ("Editors", "Atto HTML editor", "/lib/editor/atto", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Editors", "TinyMCE editor", "/lib/editor/tiny", "2022112800", "2024100701", "Standard", "To be upgraded"),
    ("Editors", "tinymce (legacy)", "", "2022112800", "", "Standard", "To be deleted"),
    # TinyMCE new plugins
    ("TinyMCE plugins", "AI placement", "/lib/editor/tiny/plugins/aiplacement", "", "2024100700", "Standard", "To be installed"),
    ("TinyMCE plugins", "HTML", "/lib/editor/tiny/plugins/html", "", "2024100700", "Standard", "To be installed"),
    ("TinyMCE plugins", "Link", "/lib/editor/tiny/plugins/link", "", "2024100700", "Standard", "To be installed"),
    ("TinyMCE plugins", "No auto-link", "/lib/editor/tiny/plugins/noautolink", "", "2024100700", "Standard", "To be installed"),
    ("TinyMCE plugins", "TinyMCE Premium", "/lib/editor/tiny/plugins/premium", "", "2024100700", "Standard", "To be installed"),
    # Enrolment
    ("Enrolment", "Manual enrolments", "/enrol/manual", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Enrolment", "Self enrolment", "/enrol/self", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Enrolment", "Guest access", "/enrol/guest", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Enrolment", "Cohort sync", "/enrol/cohort", "2022112800", "2024100700", "Standard", "To be upgraded"),
    # Auth
    ("Authentication", "Manual accounts", "/auth/manual", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Authentication", "Email self-registration", "/auth/email", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Authentication", "OAuth 2", "/auth/oauth2", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Authentication", "Web services", "/auth/webservice", "2022112800", "2024100700", "Standard", "To be upgraded"),
    # Admin tools
    ("Admin tools", "Certificate manager", "/admin/tool/certificate", "2021020800.5", "2025031804", "Additional", "To be upgraded"),
    ("Admin tools", "Multi-factor authentication", "/admin/tool/mfa", "", "2024100700", "Standard", "To be installed"),
    ("Admin tools", "TCPDF Fonts", "/admin/tool/tcpdffonts", "2021090103", "2025120101", "Additional", "To be upgraded"),
    ("Admin tools", "Learning plans", "/admin/tool/lp", "2022112800", "2024100701", "Standard", "To be upgraded"),
    ("Admin tools", "innodb (legacy)", "", "2022112800", "", "Standard", "To be deleted"),
    # MFA Factors (all new)
    ("MFA Factors", "Non-administrator", "/admin/tool/mfa/factor/admin", "", "2024100700", "Standard", "To be installed"),
    ("MFA Factors", "Email", "/admin/tool/mfa/factor/email", "", "2024100700", "Standard", "To be installed"),
    ("MFA Factors", "Authenticator app", "/admin/tool/mfa/factor/totp", "", "2024100700", "Standard", "To be installed"),
    ("MFA Factors", "Security key", "/admin/tool/mfa/factor/webauthn", "", "2024100700", "Standard", "To be installed"),
    ("MFA Factors", "SMS mobile phone", "/admin/tool/mfa/factor/sms", "", "2024100702", "Standard", "To be installed"),
    # AI (all new in 4.5)
    ("AI Placements", "Course assistance", "/ai/placement/courseassist", "", "2024100700", "Standard", "To be installed"),
    ("AI Placements", "Text editor", "/ai/placement/editor", "", "2024100700", "Standard", "To be installed"),
    ("AI Providers", "Azure AI API", "/ai/provider/azureai", "", "2024100700", "Standard", "To be installed"),
    ("AI Providers", "OpenAI API", "/ai/provider/openai", "", "2024100700", "Standard", "To be installed"),
    # Communication (new in 4.5)
    ("Communication", "Custom link", "/communication/provider/customlink", "", "2024100700", "Standard", "To be installed"),
    ("Communication", "Matrix", "/communication/provider/matrix", "", "2024100700", "Standard", "To be installed"),
    # Certificate elements
    ("Certificate elements", "Dynamic fields", "/admin/tool/certificate/element/program", "", "2025031804", "Additional", "To be installed"),
    ("Certificate elements", "modulename (legacy)", "", "2017111300", "", "Additional", "Missing from disk!"),
    # Reports
    ("Reports", "Theme usage", "/report/themeusage", "", "2024100700", "Standard", "To be installed"),
    # Custom fields (new)
    ("Custom fields", "Number", "/customfield/field/number", "", "2024100703", "Standard", "To be installed"),
    # H5P
    ("H5P frameworks", "H5P framework v1.27", "/h5p/h5plib/v127", "", "2024100700", "Standard", "To be installed"),
    ("H5P frameworks", "v124 (legacy)", "", "2022112800", "", "Standard", "To be deleted"),
    # Themes
    ("Themes", "Epsilon (airpayux)", "/theme/airpayux", "", "2026040500", "Additional (Custom)", "To be installed"),
    ("Themes", "Boost", "/theme/boost", "2022112800", "2024100700", "Standard", "To be upgraded"),
    ("Themes", "Classic", "/theme/classic", "2022112800", "2024100700", "Standard", "To be upgraded"),
    # Local plugins (custom)
    ("Local plugins", "Airpay Integrations Hub", "/local/airpay_integrations", "", "2026040500", "Additional (Custom)", "To be installed"),
    ("Local plugins", "Airpay Employee Lifecycle", "/local/airpay_lifecycle", "", "2026040500", "Additional (Custom)", "To be installed"),
    ("Local plugins", "Airpay Pages", "/local/airpay_pages", "", "2026040400", "Additional (Custom)", "To be installed"),
    ("Local plugins", "Course recompletion", "/local/recompletion", "2023012600", "2025041400", "Additional", "To be upgraded"),
    # SMS (new)
    ("SMS gateways", "AWS", "/sms/gateway/aws", "", "2024100700", "Standard", "To be installed"),
    # Payment
    ("Payment gateways", "PayPal", "/payment/gateway/paypal", "2022112800", "2024100700", "Standard", "To be upgraded"),
    # Deleted legacy plugins
    ("Legacy (deleted)", "memcached cache store", "", "2022112800", "", "Standard", "To be deleted"),
    ("Legacy (deleted)", "mongodb cache store", "", "2022112801", "", "Standard", "To be deleted"),
    ("Legacy (deleted)", "legacy log store", "", "2022112800", "", "Standard", "To be deleted"),
    ("Legacy (deleted)", "assignment offline", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "assignment online", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "assignment upload", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "assignment uploadsingle", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce ctrlhelp", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce managefiles", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce moodleemoticon", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce moodleimage", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce moodlemedia", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce moodlenolink", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce pdw", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce spellchecker", "", "2022112800", "", "Additional", "To be deleted"),
    ("Legacy (deleted)", "tinymce wrap", "", "2022112800", "", "Additional", "To be deleted"),
]

wb = Workbook()

# Summary sheet
ws = wb.active
ws.title = "Summary"
ws.append(["AIRPAY ACADEMY — Production Import Upgrade Log"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="0066A7")
ws.append(["Date: April 2026 | Source: airpayprod 6th April backup → Local XAMPP Moodle 4.5.10"])
ws.append([])

# Count by status
status_counts = {}
for p in plugins:
    s = p[6]
    status_counts[s] = status_counts.get(s, 0) + 1

ws.append(["Status", "Count", "Description"])
for col in range(1, 4):
    ws.cell(row=4, column=col).font = header_font
    ws.cell(row=4, column=col).fill = header_fill

ws.append(["To be upgraded", status_counts.get("To be upgraded", 0), "Existing plugins updated from 4.1 to 4.5"])
ws.append(["To be installed", status_counts.get("To be installed", 0), "New plugins in Moodle 4.5 (not in production 4.1)"])
ws.append(["To be deleted", status_counts.get("To be deleted", 0), "Legacy plugins removed in Moodle 4.5"])
ws.append(["Missing from disk!", status_counts.get("Missing from disk!", 0), "Plugin in DB but not on disk — needs attention"])
ws.append([])

# Count by category
ws.append(["Category", "Count"])
for col in range(1, 3):
    ws.cell(row=ws.max_row, column=col).font = header_font
    ws.cell(row=ws.max_row, column=col).fill = header_fill

cat_counts = {}
for p in plugins:
    cat_counts[p[0]] = cat_counts.get(p[0], 0) + 1
for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
    ws.append([cat, cnt])

ws.append([])
ws.append(["Custom Airpay plugins being installed:"])
for p in plugins:
    if "Custom" in p[5]:
        ws.append([f"  {p[1]}", p[2], p[4]])

# Auto-width
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

# Detail sheet
ws2 = wb.create_sheet("All Plugins")
headers = ["Category", "Plugin Name", "Directory", "Current Version", "New Version", "Source", "Status"]
ws2.append(headers)
for col in range(1, len(headers) + 1):
    ws2.cell(row=1, column=col).font = header_font
    ws2.cell(row=1, column=col).fill = header_fill
    ws2.cell(row=1, column=col).alignment = Alignment(horizontal='center')

for p in plugins:
    row_data = list(p)
    ws2.append(row_data)
    row_num = ws2.max_row
    status = p[6]
    fill = None
    if status == "To be installed":
        fill = install_fill
    elif status == "To be upgraded":
        fill = upgrade_fill
    elif status == "To be deleted":
        fill = delete_fill
    elif status == "Missing from disk!":
        fill = missing_fill
    if "Custom" in p[5]:
        fill = custom_fill
    if fill:
        for col in range(1, len(headers) + 1):
            ws2.cell(row=row_num, column=col).fill = fill
    for col in range(1, len(headers) + 1):
        ws2.cell(row=row_num, column=col).border = thin_border

for col in ws2.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Total plugins: {len(plugins)}")
print(f"Upgraded: {status_counts.get('To be upgraded', 0)}")
print(f"Installed: {status_counts.get('To be installed', 0)}")
print(f"Deleted: {status_counts.get('To be deleted', 0)}")
print(f"Missing: {status_counts.get('Missing from disk!', 0)}")
