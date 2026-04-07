"""
Extract user data from production SQL dump and generate Excel verification workbook.
Read-only analysis — no production modifications.

Output: moodle-enhancement/Production-Data-Verification.xlsx
"""
import re
import csv
import os
from datetime import datetime, timezone
from collections import Counter

# Paths
SQL_FILE = r"D:\Claude Local\airpay-ld-os\backups\Production Database backup 6th April\airpayprod-6-4.sql"
OUTPUT_XLSX = r"D:\Claude Local\airpay-ld-os\moodle-enhancement\Production-Data-Verification.xlsx"

# Column indices in mdl_user (0-based, from CREATE TABLE)
COL = {
    'id': 0, 'auth': 1, 'confirmed': 2, 'deleted': 4, 'suspended': 5,
    'username': 7, 'idnumber': 9, 'firstname': 10, 'lastname': 11,
    'email': 12, 'phone1': 14, 'department': 17, 'firstaccess': 26,
    'lastaccess': 27, 'lastlogin': 28, 'timecreated': 39,
    'open_path': 48, 'open_employeeid': 50, 'open_designation': 52,
    'open_jobfunction': 54, 'open_group': 55, 'open_location': 57,
    'open_team': 58, 'open_band': 61, 'open_hrmsrole': 62,
    'open_joindate': 72, 'open_employmenttype': 75,
}


def parse_sql_values(line):
    """Parse SQL INSERT VALUES into list of tuples, handling quoted strings with commas."""
    records = []
    # Remove the INSERT INTO ... VALUES prefix
    match = re.match(r"INSERT INTO `[^`]+` VALUES\s*", line)
    if not match:
        return records

    data = line[match.end():]
    # State machine parser
    current_record = []
    current_field = ''
    in_quote = False
    escape_next = False
    paren_depth = 0

    i = 0
    while i < len(data):
        ch = data[i]

        if escape_next:
            current_field += ch
            escape_next = False
            i += 1
            continue

        if ch == '\\':
            escape_next = True
            current_field += ch
            i += 1
            continue

        if ch == "'" and not in_quote:
            in_quote = True
            i += 1
            continue
        elif ch == "'" and in_quote:
            # Check for escaped quote ''
            if i + 1 < len(data) and data[i + 1] == "'":
                current_field += "'"
                i += 2
                continue
            in_quote = False
            i += 1
            continue

        if in_quote:
            current_field += ch
            i += 1
            continue

        # Not in quotes
        if ch == '(':
            paren_depth += 1
            if paren_depth == 1:
                current_field = ''
                current_record = []
            i += 1
            continue
        elif ch == ')':
            # Save last field
            current_record.append(current_field.strip() if current_field.strip() != 'NULL' else None)
            current_field = ''
            records.append(tuple(current_record))
            paren_depth = 0
            i += 1
            continue
        elif ch == ',' and paren_depth == 1:
            current_record.append(current_field.strip() if current_field.strip() != 'NULL' else None)
            current_field = ''
            i += 1
            continue
        elif ch == ',' and paren_depth == 0:
            # Between records
            i += 1
            continue
        elif ch == ';':
            break
        else:
            if paren_depth == 1:
                current_field += ch
            i += 1
            continue

    return records


def ts_to_date(ts_str):
    """Convert Unix timestamp string to readable date."""
    if not ts_str or ts_str == '0':
        return 'Never'
    try:
        ts = int(ts_str)
        if ts == 0:
            return 'Never'
        return datetime.fromtimestamp(ts, tz=timezone.utc).strftime('%Y-%m-%d')
    except (ValueError, OSError):
        return 'Invalid'


def clean_for_excel(val):
    """Remove illegal XML characters that openpyxl rejects."""
    if val is None:
        return None
    import re as _re
    # Remove control chars except tab, newline, carriage return
    return _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(val))


def get_field(record, field_name):
    """Safely get a field from a record tuple."""
    idx = COL.get(field_name)
    if idx is None or idx >= len(record):
        return None
    val = record[idx]
    if val is None:
        return None
    return clean_for_excel(str(val).strip())


def main():
    print("Reading SQL dump (this may take a moment for the 3.5GB file)...")

    # Step 1: Extract costcenter paths
    print("Extracting costcenter paths...")
    valid_paths = set()
    costcenter_names = {}

    with open(SQL_FILE, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            if line.startswith("INSERT INTO `mdl_local_costcenter`"):
                records = parse_sql_values(line)
                for r in records:
                    if len(r) >= 10:
                        cc_id = r[0]
                        cc_name = r[1]
                        cc_path = r[9]
                        if cc_path:
                            valid_paths.add(cc_path)
                            costcenter_names[cc_path] = cc_name
                break  # Only one INSERT line per table

    print(f"  Found {len(valid_paths)} valid costcenter paths")

    # Check: do /80 or /100 appear as valid paths?
    print(f"  Path /80 in costcenter table: {'/80' in valid_paths}")
    print(f"  Path /100 in costcenter table: {'/100' in valid_paths}")
    print(f"  Sample paths: {sorted(list(valid_paths))[:10]}")

    # Build valid path set: a user path is valid if it matches any costcenter path
    # OR if it's a deeper path under a valid costcenter (e.g., /1/2/3/13/200 is valid if /1/2/3/13 exists)
    # Orphaned = user path's top-level org doesn't exist at all
    all_path_prefixes = set(valid_paths)  # Exact paths only
    # Also add parent prefixes for hierarchical matching
    for p in valid_paths:
        parts = p.strip('/').split('/')
        for i in range(len(parts)):
            all_path_prefixes.add('/' + '/'.join(parts[:i + 1]))

    # Step 2: Extract user data
    print("Extracting user data...")
    users = []

    with open(SQL_FILE, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            if line.startswith("INSERT INTO `mdl_user`"):
                records = parse_sql_values(line)
                for r in records:
                    users.append(r)
                # There may be multiple INSERT lines
            elif users and line.startswith("INSERT INTO") and not line.startswith("INSERT INTO `mdl_user`"):
                # Moved past user table
                if len(users) > 100:
                    break

    print(f"  Found {len(users)} user records")

    # Step 3: Classify users
    orphaned = []
    duplicates_map = {}  # email -> list of users
    never_logged = []
    all_active = []
    suspended_users = []
    deleted_users = []

    for u in users:
        uid = get_field(u, 'id')
        deleted = get_field(u, 'deleted')
        suspended_flag = get_field(u, 'suspended')
        username = get_field(u, 'username')
        firstname = get_field(u, 'firstname')
        lastname = get_field(u, 'lastname')
        email = get_field(u, 'email')
        open_path = get_field(u, 'open_path')
        lastaccess = get_field(u, 'lastaccess')
        lastlogin = get_field(u, 'lastlogin')
        timecreated = get_field(u, 'timecreated')
        designation = get_field(u, 'open_designation')
        employeeid = get_field(u, 'open_employeeid')
        jobfunction = get_field(u, 'open_jobfunction')
        group = get_field(u, 'open_group')
        location = get_field(u, 'open_location')
        team = get_field(u, 'open_team')
        band = get_field(u, 'open_band')
        joindate = get_field(u, 'open_joindate')
        emptype = get_field(u, 'open_employmenttype')

        # Skip admin and guest
        if uid in ('1', '0'):
            continue

        is_deleted = (deleted == '1')
        is_suspended = (suspended_flag == '1')

        if is_deleted:
            deleted_users.append(u)
        if is_suspended and not is_deleted:
            suspended_users.append(u)

        # Check for orphaned path REGARDLESS of deleted/suspended status
        # Admin needs to know about ALL users on invalid paths

        # Resolve org name from path
        org_name = ''
        if open_path:
            # Get top-level org
            parts = open_path.strip('/').split('/')
            top_path = '/' + parts[0] if parts[0] else ''
            org_name = costcenter_names.get(top_path, '')

            # Check if user's path is valid:
            # Valid if the exact path exists, or if ANY prefix of the path exists in costcenter
            path_is_valid = False
            if open_path in all_path_prefixes:
                path_is_valid = True
            else:
                # Check if any prefix of the user's path is a valid costcenter
                for i in range(len(parts), 0, -1):
                    prefix = '/' + '/'.join(parts[:i])
                    if prefix in valid_paths:
                        path_is_valid = True
                        break

            if not path_is_valid:
                status = 'Deleted' if is_deleted else ('Suspended' if is_suspended else 'Active')
                orphaned.append({
                    'id': uid, 'username': username, 'firstname': firstname,
                    'lastname': lastname, 'email': email, 'open_path': open_path,
                    'org_name': org_name, 'designation': designation,
                    'employee_id': employeeid, 'last_access': ts_to_date(lastaccess),
                    'last_login': ts_to_date(lastlogin), 'created': ts_to_date(timecreated),
                    'jobfunction': jobfunction, 'location': location, 'team': team,
                    'band': band, 'join_date': joindate, 'employment_type': emptype,
                    'current_status': status,
                    'action': '',  # Admin fills this in
                    'new_path': '',  # Admin fills this in
                })
                if is_deleted or is_suspended:
                    continue  # Don't add to active list

        # Skip deleted/suspended users from active tracking
        if is_deleted or is_suspended:
            continue

        # Track all active users
        user_info = {
            'id': uid, 'username': username, 'firstname': firstname,
            'lastname': lastname, 'email': email, 'open_path': open_path,
            'org_name': org_name, 'designation': designation,
            'employee_id': employeeid, 'last_access': ts_to_date(lastaccess),
            'last_login': ts_to_date(lastlogin), 'created': ts_to_date(timecreated),
            'jobfunction': jobfunction, 'group': group, 'location': location,
            'team': team, 'band': band, 'join_date': joindate, 'employment_type': emptype,
        }
        all_active.append(user_info)

        # Track emails for duplicates (skip empty/placeholder emails)
        if email and '@' in email and email.lower().strip() not in ('', 'noemail@example.com'):
            email_lower = email.lower().strip()
            if email_lower not in duplicates_map:
                duplicates_map[email_lower] = []
            duplicates_map[email_lower].append(user_info)

        # Never logged in
        if (lastlogin == '0' or lastlogin is None) and (lastaccess == '0' or lastaccess is None):
            never_logged.append(user_info)

    # Filter to actual duplicates (2+ users same email)
    duplicate_emails = {k: v for k, v in duplicates_map.items() if len(v) > 1}

    # Step 4: Path distribution summary
    path_counts = Counter()
    for u in all_active:
        path = u.get('open_path', 'NULL')
        if path:
            parts = path.strip('/').split('/')
            top = '/' + parts[0] if parts[0] else 'EMPTY'
        else:
            top = 'NULL'
        path_counts[top] += 1

    # Debug: show distribution of orphaned paths
    orphan_paths = Counter(u.get('open_path', 'NULL') for u in orphaned)
    print(f"\n  Orphaned path distribution: {dict(orphan_paths)}")

    # Debug: show sample open_path values
    sample_paths = Counter(u.get('open_path', 'NULL') for u in all_active)
    print(f"  Top 10 user paths: {sample_paths.most_common(10)}")

    print(f"\n=== SUMMARY ===")
    print(f"  Total users (non-deleted): {len(all_active)}")
    print(f"  Orphaned users (invalid path): {len(orphaned)}")
    print(f"  Duplicate email addresses: {len(duplicate_emails)} emails ({sum(len(v) for v in duplicate_emails.values())} users)")
    print(f"  Never logged in: {len(never_logged)}")
    print(f"  Suspended: {len(suspended_users)}")
    print(f"  Deleted: {len(deleted_users)}")

    # Step 5: Write Excel
    print(f"\nWriting Excel workbook...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()

    # --- Styling ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066A7", end_color="0066A7", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    danger_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    action_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, num_cols):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def auto_width(ws, max_width=40):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
                cell.border = thin_border
            ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

    # ═══════════════════════════════════════
    # Sheet 1: SUMMARY
    # ═══════════════════════════════════════
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(["AIRPAY ACADEMY — Production Data Verification"])
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14, color="0066A7")
    ws_summary.append(["Generated from airpayprod backup (6th April 2026)"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value", "Notes"])
    style_header(ws_summary, 3)
    ws_summary.append(["Total Active Users", len(all_active), "Non-deleted, non-suspended"])
    ws_summary.append(["Orphaned Users", len(orphaned), "On paths that don't exist in costcenter table"])
    ws_summary.append(["Duplicate Emails", len(duplicate_emails), f"Emails shared by {sum(len(v) for v in duplicate_emails.values())} accounts"])
    ws_summary.append(["Never Logged In", len(never_logged), "Created but never accessed the LMS"])
    ws_summary.append(["Suspended Users", len(suspended_users), "Account suspended by admin"])
    ws_summary.append(["Deleted Users", len(deleted_users), "Soft-deleted (still in DB)"])
    ws_summary.append([])

    ws_summary.append(["Tenant Distribution", "", ""])
    for path, count in sorted(path_counts.items(), key=lambda x: -x[1]):
        name = costcenter_names.get(path, 'UNKNOWN/ORPHANED')
        ws_summary.append([f"  {path}", count, name])

    ws_summary.append([])
    ws_summary.append(["ACTION REQUIRED:", "", ""])
    ws_summary.append(["1. Review 'Orphaned Users' sheet — decide Reassign/Suspend/Delete for each", "", ""])
    ws_summary.append(["2. Review 'Duplicate Emails' sheet — decide which account to keep", "", ""])
    ws_summary.append(["3. Review 'Never Logged In' sheet — decide whether to keep or suspend", "", ""])
    ws_summary.append(["4. Return this file with columns filled in", "", ""])

    auto_width(ws_summary)
    # Move header styling to row 4
    for col_idx in range(1, 4):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    # ═══════════════════════════════════════
    # Sheet 2: ORPHANED USERS
    # ═══════════════════════════════════════
    ws_orphaned = wb.create_sheet("Orphaned Users")
    headers = ["User ID", "Username", "First Name", "Last Name", "Email",
               "Current Path", "Path Org (if any)", "Designation", "Employee ID",
               "Last Access", "Last Login", "Account Created", "Job Function",
               "Location", "Team", "Band", "Join Date", "Employment Type",
               "Current Status", "ACTION (fill in)", "NEW PATH (fill in)"]
    ws_orphaned.append(headers)
    style_header(ws_orphaned, len(headers))

    for u in sorted(orphaned, key=lambda x: x['open_path'] or ''):
        row = [u['id'], u['username'], u['firstname'], u['lastname'], u['email'],
               u['open_path'], u['org_name'], u['designation'], u['employee_id'],
               u['last_access'], u['last_login'], u['created'], u['jobfunction'],
               u['location'], u['team'], u['band'], u['join_date'], u['employment_type'],
               u.get('current_status', ''), '', '']
        ws_orphaned.append(row)
        # Highlight action columns
        row_num = ws_orphaned.max_row
        ws_orphaned.cell(row=row_num, column=20).fill = action_fill
        ws_orphaned.cell(row=row_num, column=21).fill = action_fill

    # Add data validation note
    ws_orphaned.append([])
    ws_orphaned.append(["ACTION values: Reassign | Suspend | Delete | Keep"])
    ws_orphaned.append(["If Reassign: fill in NEW PATH (e.g., /1/2/7 for HR department)"])

    auto_width(ws_orphaned)

    # ═══════════════════════════════════════
    # Sheet 3: DUPLICATE EMAILS
    # ═══════════════════════════════════════
    ws_dupes = wb.create_sheet("Duplicate Emails")
    headers = ["Email", "User ID", "Username", "First Name", "Last Name",
               "Org Path", "Org Name", "Last Access", "Created",
               "ACTION (fill in)"]
    ws_dupes.append(headers)
    style_header(ws_dupes, len(headers))

    for email, user_list in sorted(duplicate_emails.items()):
        for idx, u in enumerate(user_list):
            row = [email, u['id'], u['username'], u['firstname'], u['lastname'],
                   u['open_path'], u['org_name'], u['last_access'], u['created'], '']
            ws_dupes.append(row)
            row_num = ws_dupes.max_row
            ws_dupes.cell(row=row_num, column=10).fill = action_fill
            if idx > 0:
                ws_dupes.cell(row=row_num, column=1).fill = warning_fill
        # Add separator row
        ws_dupes.append([])

    ws_dupes.append(["ACTION values: Keep | Suspend | Merge"])

    auto_width(ws_dupes)

    # ═══════════════════════════════════════
    # Sheet 4: NEVER LOGGED IN
    # ═══════════════════════════════════════
    ws_never = wb.create_sheet("Never Logged In")
    headers = ["User ID", "Username", "First Name", "Last Name", "Email",
               "Org Path", "Org Name", "Designation", "Employee ID",
               "Account Created", "Employment Type", "ACTION (fill in)"]
    ws_never.append(headers)
    style_header(ws_never, len(headers))

    for u in sorted(never_logged, key=lambda x: x.get('created', '')):
        row = [u['id'], u['username'], u['firstname'], u['lastname'], u['email'],
               u['open_path'], u['org_name'], u['designation'], u['employee_id'],
               u['created'], u['employment_type'], '']
        ws_never.append(row)
        row_num = ws_never.max_row
        ws_never.cell(row=row_num, column=12).fill = action_fill

    ws_never.append([])
    ws_never.append(["ACTION values: Keep | Suspend | Delete"])

    auto_width(ws_never)

    # ═══════════════════════════════════════
    # Sheet 5: ALL ACTIVE USERS (reference)
    # ═══════════════════════════════════════
    ws_all = wb.create_sheet("All Active Users")
    headers = ["User ID", "Username", "First Name", "Last Name", "Email",
               "Org Path", "Org Name", "Designation", "Employee ID",
               "Last Access", "Last Login", "Created", "Job Function",
               "Group", "Location", "Team", "Band", "Join Date", "Employment Type"]
    ws_all.append(headers)
    style_header(ws_all, len(headers))

    for u in sorted(all_active, key=lambda x: x.get('open_path', '') or ''):
        row = [u['id'], u['username'], u['firstname'], u['lastname'], u['email'],
               u['open_path'], u['org_name'], u['designation'], u['employee_id'],
               u['last_access'], u['last_login'], u['created'], u['jobfunction'],
               u['group'], u['location'], u['team'], u['band'], u['join_date'],
               u['employment_type']]
        ws_all.append(row)

    auto_width(ws_all)

    # ═══════════════════════════════════════
    # Sheet 6: COSTCENTER MAP (reference)
    # ═══════════════════════════════════════
    ws_cc = wb.create_sheet("Costcenter Map")
    ws_cc.append(["Path", "Org Name", "User Count", "Status"])
    style_header(ws_cc, 4)

    for path in sorted(valid_paths):
        name = costcenter_names.get(path, '')
        # Count users on this exact path
        count = sum(1 for u in all_active if u.get('open_path') == path)
        ws_cc.append([path, name, count, "Active" if count > 0 else "Empty"])

    auto_width(ws_cc)

    # Save
    wb.save(OUTPUT_XLSX)
    print(f"\nExcel saved: {OUTPUT_XLSX}")
    print(f"Sheets: Summary, Orphaned Users ({len(orphaned)}), Duplicate Emails ({len(duplicate_emails)}), "
          f"Never Logged In ({len(never_logged)}), All Active Users ({len(all_active)}), Costcenter Map ({len(valid_paths)})")


if __name__ == '__main__':
    main()
